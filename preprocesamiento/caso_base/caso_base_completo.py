"""
================================================================================
CASO BASE / POLÍTICA MIOPE - Asignación de tablas quirúrgicas
ICS2122 - Taller de Investigación Operativa - Grupo 11 - Entrega 2
================================================================================

SCRIPT TODO-EN-UNO:
  1. Ejecuta la heurística greedy de asignación.
  2. Calcula KPIs.
  3. Genera Excel con resultados detallados.
  4. Genera 4 figuras (.png) para el informe.

Todos los outputs se guardan en la carpeta 'outputs/' creada al lado de este
script.

Lógica de la heurística:
  1. Ordenar lista de espera por prioridad (1 = mayor) y, en empate, por
     duración ascendente (criterio Shortest Processing Time - SPT,
     ref. Saadouli et al., 2015).
  2. Recorrer paciente por paciente, intentando asignarlo al primer
     (día, pabellón, hora) factible:
       - Pabellón disponible 8:00-18:00.
       - Respeto de la ventana horaria del servicio.
       - Cama disponible (UCI para Vascular/AV-Fistula, hospitalización
         común para el resto).
  3. Si no cabe en ningún slot del horizonte -> NO ASIGNADO.

Requisitos: pandas, numpy, openpyxl, matplotlib
  pip install pandas numpy openpyxl matplotlib
"""

import os
from collections import defaultdict

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================================
# 1. PARÁMETROS — AJUSTAR AQUÍ SI ES NECESARIO
# ============================================================================

# Ruta al Excel de datos. Cambiar si el archivo está en otra ubicación.
ARCHIVO_DATOS = os.path.join("preprocesamiento","Datos", "Datos Operaciones y lista de espera.xlsx")

# Carpeta de salida (se crea automáticamente si no existe)
CARPETA_OUTPUTS = 'outputs'

# Convención de prioridad: True si 5 = mayor, False si 1 = mayor.
PRIORIDAD_MAX_ES_MAYOR = False   # 1 = mayor prioridad

# Parámetros del problema
DIAS = 7                                # horizonte: 1 semana hábil
PABELLONES = list(range(1, 9))          # 8 pabellones
HORA_INICIO = 8 * 60                    # 8:00 en minutos
HORA_FIN = 18 * 60                      # 18:00 en minutos
CAMAS_HOSPITALIZACION = 75
CAMAS_UCI = 2

# Ventanas horarias por servicio (en minutos desde medianoche)
RESTRICCIONES_HORARIAS = {
    'ENT':           [(10*60, 15*60)],
    'General':       [(HORA_INICIO, HORA_FIN)],
    'OBGYN':         [(8*60, 13*60)],
    'Ophthalmology': [(8*60, 13*60)],
    'Orthopedics':   [(8*60, 11*60), (14*60, 18*60)],
    'Pediatrics':    [(8*60, 11*60), (14*60, 18*60)],
    'Plastic':       [(11*60, 16*60)],
    'Podiatry':      [(8*60, 13*60), (14*60, 17*60)],
    'Urology':       [(HORA_INICIO, HORA_FIN)],
    'Vascular':      [(8*60, 11*60), (15*60, 18*60)],
}


# ============================================================================
# 2. CARGA Y PREPARACIÓN DE DATOS
# ============================================================================

def cargar_datos():
    """Carga lista de espera y estima días de permanencia desde datos base."""
    db = pd.read_excel(ARCHIVO_DATOS, sheet_name='Datos base')
    le = pd.read_excel(ARCHIVO_DATOS, sheet_name='Lista de espera').copy()

    estim = (db.groupby(['Servicio', 'Descripción'])['Días de permanencia programados']
               .mean().round().astype(int).reset_index()
               .rename(columns={'Días de permanencia programados': 'dias_estimados'}))

    le = le.merge(estim, on=['Servicio', 'Descripción'], how='left')
    le['dias_estimados'] = le['dias_estimados'].fillna(
        le['Servicio'].map(db.groupby('Servicio')['Días de permanencia programados']
                             .mean().round().astype(int))
    ).fillna(0).astype(int)

    return le, db


# ============================================================================
# 3. UTILIDADES DE FACTIBILIDAD
# ============================================================================

def ventanas_factibles(servicio, duracion):
    """Ventanas horarias en las que cabe una cirugía de la duración dada."""
    return [(vi, vf) for vi, vf in RESTRICCIONES_HORARIAS[servicio]
            if vf - vi >= duracion]


def encontrar_hora_inicio(ocupacion_pabellon, ventanas, duracion):
    """Busca el primer hueco factible. Retorna hora inicio o None."""
    for vi, vf in ventanas:
        cursor = vi
        for ini, fin in ocupacion_pabellon:
            if fin <= cursor:
                continue
            if ini >= cursor + duracion and ini <= vf:
                return cursor
            cursor = max(cursor, fin)
            if cursor + duracion > vf:
                break
        else:
            if cursor + duracion <= vf:
                return cursor
    return None


def insertar_ordenado(lista_intervalos, intervalo):
    lista_intervalos.append(intervalo)
    lista_intervalos.sort()


# ============================================================================
# 4. HEURÍSTICA GREEDY DE ASIGNACIÓN
# ============================================================================

def asignar_greedy(lista_espera):
    """Política miope: ordena por prioridad y SPT, asigna al primer hueco."""
    df = lista_espera.copy()
    if PRIORIDAD_MAX_ES_MAYOR:
        df['_prio_orden'] = -df['Prioridad de paciente']
    else:
        df['_prio_orden'] = df['Prioridad de paciente']
    df = df.sort_values(['_prio_orden', 'Duración agendada (min)'],
                        ascending=[True, True]).reset_index(drop=True)

    ocupacion = {d: {p: [] for p in PABELLONES} for d in range(DIAS)}
    camas_uso = defaultdict(int)
    uci_uso = defaultdict(int)

    asignaciones = []
    no_asignados = []

    for _, paciente in df.iterrows():
        servicio = paciente['Servicio']
        duracion = int(paciente['Duración agendada (min)'])
        dias_perm = int(paciente['dias_estimados'])
        es_uci = (servicio == 'Vascular' and
                  'fistula' in str(paciente['Descripción']).lower())

        dias_cama = 2 if es_uci else dias_perm
        capacidad_camas = CAMAS_UCI if es_uci else CAMAS_HOSPITALIZACION
        uso_actual = uci_uso if es_uci else camas_uso

        ventanas = ventanas_factibles(servicio, duracion)
        if not ventanas:
            no_asignados.append({**paciente.to_dict(),
                                 'motivo': 'duración > ventana servicio'})
            continue

        asignado = False
        for d in range(DIAS):
            if asignado:
                break

            if dias_cama > 0:
                cabe_cama = all(
                    uso_actual[d + k] + 1 <= capacidad_camas
                    for k in range(dias_cama)
                    if d + k < DIAS
                )
                if not cabe_cama:
                    continue

            p = int(paciente['OR Suite'])
            hora = encontrar_hora_inicio(ocupacion[d][p], ventanas, duracion)
            if hora is not None:
                insertar_ordenado(ocupacion[d][p], (hora, hora + duracion))
                if dias_cama > 0:
                    for k in range(dias_cama):
                        if d + k < DIAS:
                            uso_actual[d + k] += 1
                asignaciones.append({
                    'Correlativo': paciente['Correlativo'],
                    'Servicio': servicio,
                    'Descripción': paciente['Descripción'],
                    'Prioridad': paciente['Prioridad de paciente'],
                    'Duración (min)': duracion,
                    'Días cama': dias_cama,
                    'Tipo cama': 'UCI' if es_uci else (
                        'Hosp.' if dias_cama > 0 else 'Ambulatorio'),
                    'Día': d + 1,
                    'Pabellón': p,
                    'Hora inicio (min)': hora,
                    'Hora inicio': f'{hora//60:02d}:{hora%60:02d}',
                    'Hora fin': f'{(hora+duracion)//60:02d}:{(hora+duracion)%60:02d}',
                })
                asignado = True

        if not asignado:
            no_asignados.append({**paciente.to_dict(),
                                 'motivo': 'sin capacidad pabellón/cama'})

    return (pd.DataFrame(asignaciones),
            pd.DataFrame(no_asignados),
            ocupacion, camas_uso, uci_uso)


# ============================================================================
# 5. CÁLCULO DE KPIs
# ============================================================================

def calcular_kpis(asign, no_asign, ocupacion, camas_uso, uci_uso, lista_espera):
    total_le = len(lista_espera)
    total_asig = len(asign)
    total_no = len(no_asign)

    pacientes_programados = total_asig
    tasa_atencion = total_asig / total_le if total_le else 0

    if PRIORIDAD_MAX_ES_MAYOR:
        score_prioridad = asign['Prioridad'].sum() if total_asig else 0
        score_max = lista_espera['Prioridad de paciente'].sum()
    else:
        # Convertir prioridad (1=mayor) a score donde más alto = mejor
        score_prioridad = (6 - asign['Prioridad']).sum() if total_asig else 0
        score_max = (6 - lista_espera['Prioridad de paciente']).sum()
    score_pct = score_prioridad / score_max if score_max else 0

    minutos_disp_total = DIAS * len(PABELLONES) * (HORA_FIN - HORA_INICIO)
    minutos_usados = 0
    for d in ocupacion:
        for p in ocupacion[d]:
            for ini, fin in ocupacion[d][p]:
                minutos_usados += (fin - ini)
    utilizacion_pabellones = minutos_usados / minutos_disp_total

    cap_camas_total = DIAS * CAMAS_HOSPITALIZACION
    uso_camas_total = sum(camas_uso.get(d, 0) for d in range(DIAS))
    utilizacion_camas = uso_camas_total / cap_camas_total

    cap_uci_total = DIAS * CAMAS_UCI
    uso_uci_total = sum(uci_uso.get(d, 0) for d in range(DIAS))
    utilizacion_uci = uso_uci_total / cap_uci_total

    holgura_total = 0
    for d in ocupacion:
        for p in ocupacion[d]:
            if ocupacion[d][p]:
                ultimo_fin = max(fin for _, fin in ocupacion[d][p])
                holgura_total += max(0, HORA_FIN - ultimo_fin)
            else:
                holgura_total += (HORA_FIN - HORA_INICIO)
    holgura_promedio = holgura_total / (DIAS * len(PABELLONES))

    if total_asig:
        por_prio = asign.groupby('Prioridad').size().reset_index(name='asignados')
    else:
        por_prio = pd.DataFrame(columns=['Prioridad', 'asignados'])
    demanda_prio = (lista_espera.groupby('Prioridad de paciente').size()
                    .reset_index(name='demanda')
                    .rename(columns={'Prioridad de paciente': 'Prioridad'}))
    resumen_prio = demanda_prio.merge(por_prio, on='Prioridad', how='left').fillna(0)
    resumen_prio['asignados'] = resumen_prio['asignados'].astype(int)
    resumen_prio['tasa_atencion'] = (resumen_prio['asignados']
                                     / resumen_prio['demanda']).round(3)

    if total_asig:
        por_serv = asign.groupby('Servicio').size().reset_index(name='asignados')
    else:
        por_serv = pd.DataFrame(columns=['Servicio', 'asignados'])
    demanda_serv = (lista_espera.groupby('Servicio').size()
                    .reset_index(name='demanda'))
    resumen_serv = demanda_serv.merge(por_serv, on='Servicio', how='left').fillna(0)
    resumen_serv['asignados'] = resumen_serv['asignados'].astype(int)
    resumen_serv['tasa_atencion'] = (resumen_serv['asignados']
                                     / resumen_serv['demanda']).round(3)

    kpis = {
        'pacientes_en_lista_espera': total_le,
        'pacientes_programados': pacientes_programados,
        'pacientes_no_asignados': total_no,
        'tasa_atencion': round(tasa_atencion, 4),
        'score_prioridad_obtenido': int(score_prioridad),
        'score_prioridad_maximo': int(score_max),
        'score_prioridad_pct': round(score_pct, 4),
        'utilizacion_pabellones': round(utilizacion_pabellones, 4),
        'utilizacion_camas_hosp': round(utilizacion_camas, 4),
        'utilizacion_uci': round(utilizacion_uci, 4),
        'holgura_promedio_min': round(holgura_promedio, 1),
        'minutos_pabellon_usados': minutos_usados,
        'minutos_pabellon_disponibles': minutos_disp_total,
    }

    return kpis, resumen_prio, resumen_serv


# ============================================================================
# 6. EXPORTAR EXCEL DE RESULTADOS
# ============================================================================

def exportar_excel(asign, no_asign, kpis, res_prio, res_serv, ruta_salida):
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', name='Arial')
    header_fill = PatternFill('solid', start_color='1F4E78')

    # Hoja KPIs
    ws = wb.active
    ws.title = 'KPIs'
    ws['A1'] = 'KPIs - Caso Base (Política Miope Greedy)'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')
    ws.merge_cells('A1:C1')
    ws.append([])
    ws.append(['Indicador', 'Valor', 'Comentario'])
    for c in ws[3]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    convencion = '5=mayor' if PRIORIDAD_MAX_ES_MAYOR else '1=mayor'
    filas_kpi = [
        ('Pacientes en lista de espera', kpis['pacientes_en_lista_espera'], 'Demanda total'),
        ('Pacientes programados', kpis['pacientes_programados'], 'KPI - Capacidad de atención'),
        ('Pacientes no asignados', kpis['pacientes_no_asignados'], ''),
        ('Tasa de atención (%)', f"{kpis['tasa_atencion']*100:.2f}%", ''),
        ('Score de prioridad obtenido', kpis['score_prioridad_obtenido'],
         f'Convención prioridad: {convencion}'),
        ('Score de prioridad máximo', kpis['score_prioridad_maximo'], 'Si se atendiera toda la lista'),
        ('Score de prioridad (%)', f"{kpis['score_prioridad_pct']*100:.2f}%", ''),
        ('Utilización de pabellones (%)', f"{kpis['utilizacion_pabellones']*100:.2f}%", ''),
        ('Utilización camas hospitalización (%)', f"{kpis['utilizacion_camas_hosp']*100:.2f}%", ''),
        ('Utilización UCI (%)', f"{kpis['utilizacion_uci']*100:.2f}%", 'AV-Fistula (Vascular)'),
        ('Holgura promedio por pabellón (min/día)', kpis['holgura_promedio_min'], ''),
        ('Minutos pabellón usados', kpis['minutos_pabellon_usados'], ''),
        ('Minutos pabellón disponibles', kpis['minutos_pabellon_disponibles'],
         f'{DIAS} días x {len(PABELLONES)} pab x {HORA_FIN-HORA_INICIO} min'),
    ]
    for fila in filas_kpi:
        ws.append(list(fila))
    for col, w in zip('ABC', [42, 22, 50]):
        ws.column_dimensions[col].width = w

    # Por prioridad
    ws2 = wb.create_sheet('Por prioridad')
    ws2.append(['Resumen por prioridad'])
    ws2['A1'].font = Font(bold=True, size=12, name='Arial')
    ws2.append([])
    ws2.append(['Prioridad', 'Demanda', 'Asignados', 'Tasa atención'])
    for c in ws2[3]:
        c.font = header_font
        c.fill = header_fill
    for _, row in res_prio.iterrows():
        ws2.append([int(row['Prioridad']), int(row['demanda']),
                    int(row['asignados']), f"{row['tasa_atencion']*100:.1f}%"])
    for col, w in zip('ABCD', [12, 12, 12, 18]):
        ws2.column_dimensions[col].width = w

    # Por servicio
    ws3 = wb.create_sheet('Por servicio')
    ws3.append(['Resumen por servicio'])
    ws3['A1'].font = Font(bold=True, size=12, name='Arial')
    ws3.append([])
    ws3.append(['Servicio', 'Demanda', 'Asignados', 'Tasa atención'])
    for c in ws3[3]:
        c.font = header_font
        c.fill = header_fill
    for _, row in res_serv.sort_values('demanda', ascending=False).iterrows():
        ws3.append([row['Servicio'], int(row['demanda']),
                    int(row['asignados']), f"{row['tasa_atencion']*100:.1f}%"])
    for col, w in zip('ABCD', [16, 12, 12, 18]):
        ws3.column_dimensions[col].width = w

    # Asignaciones
    ws4 = wb.create_sheet('Asignaciones')
    ws4.append(['Tabla quirúrgica generada por el caso base'])
    ws4['A1'].font = Font(bold=True, size=12, name='Arial')
    ws4.append([])
    cols = ['Correlativo', 'Servicio', 'Descripción', 'Prioridad', 'Duración (min)',
            'Días cama', 'Tipo cama', 'Día', 'Pabellón', 'Hora inicio', 'Hora fin']
    ws4.append(cols)
    for c in ws4[3]:
        c.font = header_font
        c.fill = header_fill
    for _, row in asign.sort_values(['Día', 'Pabellón', 'Hora inicio (min)']).iterrows():
        ws4.append([row[c] for c in cols])
    for col, w in zip('ABCDEFGHIJK', [12, 14, 38, 11, 14, 11, 13, 7, 10, 12, 12]):
        ws4.column_dimensions[col].width = w

    # No asignados
    ws5 = wb.create_sheet('No asignados')
    ws5.append(['Pacientes no asignados'])
    ws5['A1'].font = Font(bold=True, size=12, name='Arial')
    ws5.append([])
    cols_na = ['Correlativo', 'Servicio', 'Descripción',
               'Duración agendada (min)', 'Prioridad de paciente', 'motivo']
    ws5.append(cols_na)
    for c in ws5[3]:
        c.font = header_font
        c.fill = header_fill
    for _, row in no_asign.iterrows():
        ws5.append([row.get(c, '') for c in cols_na])
    for col, w in zip('ABCDEF', [12, 14, 38, 18, 18, 30]):
        ws5.column_dimensions[col].width = w

    wb.save(ruta_salida)

    # =========================
    # Exportar CSV de asignaciones
    # =========================
    ruta_csv = ruta_salida.replace('.xlsx', '_asignaciones.csv')

    cols = ['Correlativo', 'Servicio', 'Descripción', 'Prioridad', 'Duración (min)',
            'Días cama', 'Tipo cama', 'Día', 'Pabellón', 'Hora inicio', 'Hora fin']

    asign_export = (
        asign
        .sort_values(['Día', 'Pabellón', 'Hora inicio (min)'])
        [cols]
    )

    asign_export.to_csv(ruta_csv, index=False, encoding='utf-8-sig')


# ============================================================================
# 7. GENERAR FIGURAS
# ============================================================================

def generar_figuras(asign, ocup, camas, uci, res_prio, carpeta):
    plt.rcParams.update({'font.family': 'serif', 'font.size': 10})

    # Fig 1: atención por prioridad
    fig, ax = plt.subplots(figsize=(7, 4))
    prio_orden = res_prio.sort_values('Prioridad')
    x = prio_orden['Prioridad'].astype(int).astype(str)
    ax.bar(x, prio_orden['demanda'], color='#A6C4DE', label='Demanda', edgecolor='black')
    ax.bar(x, prio_orden['asignados'], color='#1F4E78', label='Asignados', edgecolor='black')
    for i, (d, a) in enumerate(zip(prio_orden['demanda'], prio_orden['asignados'])):
        pct = a/d*100 if d else 0
        ax.text(i, d + 8, f'{pct:.0f}%', ha='center', fontsize=9)
    convencion = '5 = mayor' if PRIORIDAD_MAX_ES_MAYOR else '1 = mayor'
    ax.set_xlabel(f'Prioridad clínica ({convencion})')
    ax.set_ylabel('Pacientes')
    ax.set_title('Caso base: atención por nivel de prioridad')
    ax.legend()
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta, 'fig_atencion_prioridad.png'), dpi=150)
    plt.close()

    # Fig 2: utilización de pabellones por día
    util = []
    for d in range(DIAS):
        for p in PABELLONES:
            usado = sum(f - i for i, f in ocup[d][p])
            util.append({'Día': d+1, 'Pabellón': p,
                         'Utilización': usado/(HORA_FIN-HORA_INICIO)})
    util_df = pd.DataFrame(util)
    pivot = util_df.pivot(index='Pabellón', columns='Día', values='Utilización')

    fig, ax = plt.subplots(figsize=(7, 4.5))
    im = ax.imshow(pivot.values, cmap='YlOrRd', vmin=0, vmax=1, aspect='auto')
    ax.set_xticks(range(DIAS))
    ax.set_xticklabels([f'Día {d+1}' for d in range(DIAS)])
    ax.set_yticks(range(len(PABELLONES)))
    ax.set_yticklabels([f'Pab. {p}' for p in PABELLONES])
    for i in range(len(PABELLONES)):
        for j in range(DIAS):
            ax.text(j, i, f'{pivot.values[i,j]*100:.0f}%',
                    ha='center', va='center',
                    color='black' if pivot.values[i,j] < 0.6 else 'white',
                    fontsize=9)
    ax.set_title('Caso base: utilización de pabellones (%)')
    plt.colorbar(im, ax=ax, label='Utilización')
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta, 'fig_utilizacion_pabellones.png'), dpi=150)
    plt.close()

    # Fig 3: ocupación de camas
    fig, ax = plt.subplots(figsize=(7, 4))
    dias_x = [f'Día {d+1}' for d in range(DIAS)]
    camas_y = [camas.get(d, 0) for d in range(DIAS)]
    ax.bar(dias_x, camas_y, color='#1F4E78',
           label=f'Camas hospitalización (cap. {CAMAS_HOSPITALIZACION})',
           edgecolor='black')
    ax.axhline(CAMAS_HOSPITALIZACION, color='red', linestyle='--', linewidth=1,
               label='Capacidad camas hosp.')
    for i, v in enumerate(camas_y):
        ax.text(i, v + 1, str(v), ha='center', fontsize=9)
    ax.set_ylabel('Camas ocupadas')
    ax.set_title('Caso base: ocupación diaria de camas hospitalización')
    ax.legend(loc='lower right')
    ax.grid(axis='y', alpha=0.3)
    ax.set_ylim(0, CAMAS_HOSPITALIZACION + 10)
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta, 'fig_ocupacion_camas.png'), dpi=150)
    plt.close()

    # Fig 4: Gantt día 1
    fig, ax = plt.subplots(figsize=(11, 5))
    colores_serv = {
        'ENT': '#E69F00', 'General': '#56B4E9', 'OBGYN': '#009E73',
        'Ophthalmology': '#F0E442', 'Orthopedics': '#0072B2',
        'Pediatrics': '#D55E00', 'Plastic': '#CC79A7', 'Podiatry': '#999999',
        'Urology': '#882255', 'Vascular': '#117733'
    }
    asign_d1 = asign[asign['Día'] == 1]
    for _, row in asign_d1.iterrows():
        ax.barh(row['Pabellón'], row['Duración (min)'],
                left=row['Hora inicio (min)'],
                color=colores_serv[row['Servicio']],
                edgecolor='black', linewidth=0.5)
    ax.set_yticks(PABELLONES)
    ax.set_yticklabels([f'Pab. {p}' for p in PABELLONES])
    ax.set_xticks(range(8*60, 18*60+1, 60))
    ax.set_xticklabels([f'{h}:00' for h in range(8, 19)])
    ax.set_xlim(8*60-10, 18*60+10)
    ax.invert_yaxis()
    ax.set_xlabel('Hora del día')
    ax.set_title('Caso base: tabla quirúrgica - Día 1')
    ax.grid(axis='x', alpha=0.3)
    handles = [mpatches.Patch(color=c, label=s) for s, c in colores_serv.items()]
    ax.legend(handles=handles, bbox_to_anchor=(1.02, 1), loc='upper left',
              fontsize=8, title='Servicio')
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta, 'fig_gantt_dia1.png'),
                dpi=150, bbox_inches='tight')
    plt.close()


# ============================================================================
# 8. EJECUCIÓN PRINCIPAL
# ============================================================================

def main():
    print("=" * 78)
    print("CASO BASE - POLÍTICA MIOPE GREEDY")
    print("ICS2122 - Grupo 11 - Entrega 2")
    print("=" * 78)

    # Crear carpeta de outputs
    os.makedirs(CARPETA_OUTPUTS, exist_ok=True)
    print(f"\nCarpeta de salida: {os.path.abspath(CARPETA_OUTPUTS)}")

    # Cargar datos
    le, db = cargar_datos()
    print(f"\n[1] Pacientes en lista de espera: {len(le)}")
    convencion = '5 = mayor' if PRIORIDAD_MAX_ES_MAYOR else '1 = mayor'
    print(f"    Convención de prioridad: {convencion}")

    # Heurística
    print(f"\n[2] Ejecutando heurística greedy (Prioridad + SPT)...")
    asign, no_asign, ocup, camas, uci = asignar_greedy(le)
    print(f"    -> Asignados: {len(asign)}  |  No asignados: {len(no_asign)}")

    # KPIs
    print("\n[3] Cálculo de KPIs:")
    kpis, res_prio, res_serv = calcular_kpis(asign, no_asign, ocup, camas, uci, le)
    for k, v in kpis.items():
        print(f"    {k:35s}: {v}")

    print("\n[4] Atención por prioridad:")
    print(res_prio.to_string(index=False))

    print("\n[5] Atención por servicio:")
    print(res_serv.to_string(index=False))

    # Exportar
    ruta_xlsx = os.path.join(CARPETA_OUTPUTS, 'caso_base_resultados.xlsx')
    exportar_excel(asign, no_asign, kpis, res_prio, res_serv, ruta_xlsx)
    print(f"\n[6] Excel guardado: {ruta_xlsx}")

    generar_figuras(asign, ocup, camas, uci, res_prio, CARPETA_OUTPUTS)
    print(f"[7] Figuras guardadas en: {CARPETA_OUTPUTS}/")
    print(f"    - fig_atencion_prioridad.png")
    print(f"    - fig_utilizacion_pabellones.png")
    print(f"    - fig_ocupacion_camas.png")
    print(f"    - fig_gantt_dia1.png")

    print("\nListo.")
    return asign, no_asign, kpis




if __name__ == '__main__':
    main()
