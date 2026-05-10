"""
================================================================================
ANÁLISIS DE RESULTADOS - PLANIFICACIÓN MIP (7 DÍAS)
ICS2122 - Taller de Investigación Operativa - Grupo 11
================================================================================

SCRIPT TODO-EN-UNO:
  1. Lee la planificación MIP desde 7_dias.csv
  2. Calcula KPIs comparando con la lista de espera original.
  3. Genera Excel con resultados detallados.
  4. Genera 4 figuras (.png) para el informe.

Todos los outputs se guardan en la carpeta 'resultados_planificacion_MIP/'
dentro de la carpeta de este script.

Requisitos: pandas, numpy, openpyxl, matplotlib
"""

import os
from collections import defaultdict

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================================
# 1. PARÁMETROS — AJUSTAR AQUÍ SI ES NECESARIO
# ============================================================================

# Ruta al CSV de planificación MIP (relativa a este script)
ARCHIVO_PLANIFICACION = os.path.join("Simulacion", "Estado_Inicial", "7_dias.csv")

# Ruta al Excel de datos para comparar con lista de espera
ARCHIVO_DATOS = os.path.join("preprocesamiento", "Datos", "Datos Operaciones y lista de espera.xlsx")

# Carpeta de salida para los resultados (se crea automáticamente)
CARPETA_RESULTADOS = os.path.join("outputs", "resultados_planificacion_MIP")

# Convención de prioridad: True si 5 = mayor, False si 1 = mayor.
PRIORIDAD_MAX_ES_MAYOR = False   # 1 = mayor prioridad

# Parámetros del problema
DIAS = 7                                # horizonte: 1 semana (7 días)
PABELLONES = list(range(1, 9))          # 8 pabellones
HORA_INICIO = 8 * 60                    # 8:00 en minutos
HORA_FIN = 18 * 60                      # 18:00 en minutos
CAMAS_HOSPITALIZACION = 75
CAMAS_UCI = 2

# Ventanas horarias por servicio (en minutos desde medianoche)
RESTRICCIONES_HORARIAS = {
    'ENT':           [(10*60, 15*60)],
    'General':       [(HORA_INICIO, HORA_FIN)],
    'OBGYN':         [(8*60, 13*60)],
    'Ophthalmology': [(8*60, 13*60)],
    'Orthopedics':   [(8*60, 11*60), (14*60, 18*60)],
    'Pediatrics':    [(8*60, 11*60), (14*60, 18*60)],
    'Plastic':       [(11*60, 16*60)],
    'Podiatry':      [(8*60, 13*60), (14*60, 17*60)],
    'Urology':       [(HORA_INICIO, HORA_FIN)],
    'Vascular':      [(8*60, 11*60), (15*60, 18*60)],
}


# ============================================================================
# 2. CARGA Y PREPARACIÓN DE DATOS
# ============================================================================

def cargar_datos_base():
    """Carga datos base para estimaciones de permanencia."""
    db = pd.read_excel(ARCHIVO_DATOS, sheet_name='Datos base')
    return db


def cargar_lista_espera():
    """Carga y prepara la lista de espera."""
    le = pd.read_excel(ARCHIVO_DATOS, sheet_name='Lista de espera').copy()
    
    # Crear mapping de ID para buscar en el CSV
    le['Correlativo'] = range(1, len(le) + 1)  # Correlativo = índice + 1
    
    return le


def cargar_planificacion():
    """Carga la planificación desde el CSV."""
    df = pd.read_csv(ARCHIVO_PLANIFICACION)
    
    # Convertir horas de formato string a minutos
    def hora_a_minutos(hora_str):
        h, m = map(int, hora_str.split(':'))
        return h * 60 + m
    
    df['Hora inicio (min)'] = df['Hora inicio'].apply(hora_a_minutos)
    df['Hora fin (min)'] = df['Hora fin'].apply(hora_a_minutos)
    df['Duración (min)'] = df['Duración (min)']
    
    # Transformar Prioridad de vuelta a escala original (1=mayor)
    # El CSV contiene valores transformados del modelo MIP (6 - prioridad_original)
    df['Prioridad'] = 6 - df['Prioridad']
    
    # Normalizar nombres de columnas para compatibilidad
    df = df.rename(columns={
        'Correlativo': 'ID paciente',
        'Pabellón': 'Pabellón',
        'Día': 'Día',
        'Permanencia': 'Permanencia_estimada',
    })
    
    return df


# ============================================================================
# 3. ANÁLISIS DE ASIGNACIONES
# ============================================================================

def analizar_asignaciones(planif, lista_espera):
    """Identifica pacientes asignados y no asignados."""
    
    # Los asignados son los que están en el CSV
    asignados = planif.copy()
    
    # Los no asignados son los que están en la lista de espera pero no en el CSV
    ids_asignados = set(asignados['ID paciente'].unique())
    ids_lista_espera = set(lista_espera.index)
    
    ids_no_asignados = ids_lista_espera - ids_asignados
    
    no_asignados = lista_espera.loc[list(ids_no_asignados)].copy()
    no_asignados['motivo'] = 'No incluido en planificación MIP'
    
    return asignados, no_asignados


# ============================================================================
# 4. RECONSTRUCCIÓN DE OCUPACIÓN Y USO DE CAMAS
# ============================================================================

def reconstruir_ocupacion_y_camas(asignados):
    """Reconstruye la ocupación de pabellones y uso de camas desde la planificación."""
    
    ocupacion = {d: {p: [] for p in PABELLONES} for d in range(DIAS)}
    camas_uso = defaultdict(int)
    uci_uso = defaultdict(int)
    
    for _, row in asignados.iterrows():
        dia = int(row['Día']) - 1  # Convertir a índice 0-based
        pabellon = int(row['Pabellón'])
        hora_inicio = int(row['Hora inicio (min)'])
        duracion = int(row['Duración (min)'])
        
        # Registrar ocupación del pabellón
        ocupacion[dia][pabellon].append((hora_inicio, hora_inicio + duracion))
        
        # Registrar uso de camas
        permanencia = int(row['Permanencia_estimada'])
        requiere_uci = row['Requiere UCI']
        
        if permanencia > 0:
            dias_cama = 2 if requiere_uci else permanencia
            tipo_uso = uci_uso if requiere_uci else camas_uso
            
            for k in range(dias_cama):
                if dia + k < DIAS:
                    tipo_uso[dia + k] += 1
    
    # Ordenar intervalos de ocupación
    for d in ocupacion:
        for p in ocupacion[d]:
            ocupacion[d][p].sort()
    
    return ocupacion, camas_uso, uci_uso


# ============================================================================
# 5. CÁLCULO DE KPIs
# ============================================================================

def calcular_kpis(asign, no_asign, ocupacion, camas_uso, uci_uso, lista_espera):
    """Calcula KPIs de la planificación."""
    
    total_le = len(lista_espera)
    total_asig = len(asign)
    total_no = len(no_asign)

    pacientes_programados = total_asig
    tasa_atencion = total_asig / total_le if total_le else 0

    if PRIORIDAD_MAX_ES_MAYOR:
        score_prioridad = asign['Prioridad'].sum() if total_asig else 0
        score_max = lista_espera['Prioridad de paciente'].sum()
    else:
        # Convertir prioridad (1=mayor) a score donde más alto = mejor
        score_prioridad = (6 - asign['Prioridad']).sum() if total_asig else 0
        score_max = (6 - lista_espera['Prioridad de paciente']).sum()
    score_pct = score_prioridad / score_max if score_max else 0

    minutos_disp_total = DIAS * len(PABELLONES) * (HORA_FIN - HORA_INICIO)
    minutos_usados = 0
    for d in ocupacion:
        for p in ocupacion[d]:
            for ini, fin in ocupacion[d][p]:
                minutos_usados += (fin - ini)
    utilizacion_pabellones = minutos_usados / minutos_disp_total

    cap_camas_total = DIAS * CAMAS_HOSPITALIZACION
    uso_camas_total = sum(camas_uso.get(d, 0) for d in range(DIAS))
    utilizacion_camas = uso_camas_total / cap_camas_total

    cap_uci_total = DIAS * CAMAS_UCI
    uso_uci_total = sum(uci_uso.get(d, 0) for d in range(DIAS))
    utilizacion_uci = uso_uci_total / cap_uci_total

    holgura_total = 0
    for d in ocupacion:
        for p in ocupacion[d]:
            if ocupacion[d][p]:
                ultimo_fin = max(fin for _, fin in ocupacion[d][p])
                holgura_total += max(0, HORA_FIN - ultimo_fin)
            else:
                holgura_total += (HORA_FIN - HORA_INICIO)
    holgura_promedio = holgura_total / (DIAS * len(PABELLONES))

    if total_asig:
        por_prio = asign.groupby('Prioridad').size().reset_index(name='asignados')
    else:
        por_prio = pd.DataFrame(columns=['Prioridad', 'asignados'])
    demanda_prio = (lista_espera.groupby('Prioridad de paciente').size()
                    .reset_index(name='demanda')
                    .rename(columns={'Prioridad de paciente': 'Prioridad'}))
    resumen_prio = demanda_prio.merge(por_prio, on='Prioridad', how='left').fillna(0)
    resumen_prio['asignados'] = resumen_prio['asignados'].astype(int)
    resumen_prio['tasa_atencion'] = (resumen_prio['asignados']
                                     / resumen_prio['demanda']).round(3)

    if total_asig:
        por_serv = asign.groupby('Servicio').size().reset_index(name='asignados')
    else:
        por_serv = pd.DataFrame(columns=['Servicio', 'asignados'])
    demanda_serv = (lista_espera.groupby('Servicio').size()
                    .reset_index(name='demanda'))
    resumen_serv = demanda_serv.merge(por_serv, on='Servicio', how='left').fillna(0)
    resumen_serv['asignados'] = resumen_serv['asignados'].astype(int)
    resumen_serv['tasa_atencion'] = (resumen_serv['asignados']
                                     / resumen_serv['demanda']).round(3)

    kpis = {
        'pacientes_en_lista_espera': total_le,
        'pacientes_programados': pacientes_programados,
        'pacientes_no_asignados': total_no,
        'tasa_atencion': round(tasa_atencion, 4),
        'score_prioridad_obtenido': int(score_prioridad),
        'score_prioridad_maximo': int(score_max),
        'score_prioridad_pct': round(score_pct, 4),
        'utilizacion_pabellones': round(utilizacion_pabellones, 4),
        'utilizacion_camas_hosp': round(utilizacion_camas, 4),
        'utilizacion_uci': round(utilizacion_uci, 4),
        'holgura_promedio_min': round(holgura_promedio, 1),
        'minutos_pabellon_usados': minutos_usados,
        'minutos_pabellon_disponibles': minutos_disp_total,
    }

    return kpis, resumen_prio, resumen_serv


# ============================================================================
# 6. EXPORTAR EXCEL DE RESULTADOS
# ============================================================================

def exportar_excel(asign, no_asign, kpis, res_prio, res_serv, ruta_salida):
    """Exporta resultados a un archivo Excel."""
    
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', name='Arial')
    header_fill = PatternFill('solid', start_color='1F4E78')

    # Hoja KPIs
    ws = wb.active
    ws.title = 'KPIs'
    ws['A1'] = 'KPIs - Planificación MIP (7 días)'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')
    ws.merge_cells('A1:C1')
    ws.append([])
    ws.append(['Indicador', 'Valor', 'Comentario'])
    for c in ws[3]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    convencion = '5=mayor' if PRIORIDAD_MAX_ES_MAYOR else '1=mayor'
    filas_kpi = [
        ('Pacientes en lista de espera', kpis['pacientes_en_lista_espera'], 'Demanda total'),
        ('Pacientes programados', kpis['pacientes_programados'], 'KPI - Capacidad de atención'),
        ('Pacientes no asignados', kpis['pacientes_no_asignados'], ''),
        ('Tasa de atención (%)', f"{kpis['tasa_atencion']*100:.2f}%", ''),
        ('Score de prioridad obtenido', kpis['score_prioridad_obtenido'],
         f'Convención prioridad: {convencion}'),
        ('Score de prioridad máximo', kpis['score_prioridad_maximo'], 'Si se atendiera toda la lista'),
        ('Score de prioridad (%)', f"{kpis['score_prioridad_pct']*100:.2f}%", ''),
        ('Utilización de pabellones (%)', f"{kpis['utilizacion_pabellones']*100:.2f}%", ''),
        ('Utilización camas hospitalización (%)', f"{kpis['utilizacion_camas_hosp']*100:.2f}%", ''),
        ('Utilización UCI (%)', f"{kpis['utilizacion_uci']*100:.2f}%", 'AV-Fistula (Vascular)'),
        ('Holgura promedio por pabellón (min/día)', kpis['holgura_promedio_min'], ''),
        ('Minutos pabellón usados', kpis['minutos_pabellon_usados'], ''),
        ('Minutos pabellón disponibles', kpis['minutos_pabellon_disponibles'],
         f'{DIAS} días x {len(PABELLONES)} pab x {HORA_FIN-HORA_INICIO} min'),
    ]
    for fila in filas_kpi:
        ws.append(list(fila))
    for col, w in zip('ABC', [42, 22, 50]):
        ws.column_dimensions[col].width = w

    # Por prioridad
    ws2 = wb.create_sheet('Por prioridad')
    ws2.append(['Resumen por prioridad'])
    ws2['A1'].font = Font(bold=True, size=12, name='Arial')
    ws2.append([])
    ws2.append(['Prioridad', 'Demanda', 'Asignados', 'Tasa atención'])
    for c in ws2[3]:
        c.font = header_font
        c.fill = header_fill
    for _, row in res_prio.iterrows():
        ws2.append([int(row['Prioridad']), int(row['demanda']),
                    int(row['asignados']), f"{row['tasa_atencion']*100:.1f}%"])
    for col, w in zip('ABCD', [12, 12, 12, 18]):
        ws2.column_dimensions[col].width = w

    # Por servicio
    ws3 = wb.create_sheet('Por servicio')
    ws3.append(['Resumen por servicio'])
    ws3['A1'].font = Font(bold=True, size=12, name='Arial')
    ws3.append([])
    ws3.append(['Servicio', 'Demanda', 'Asignados', 'Tasa atención'])
    for c in ws3[3]:
        c.font = header_font
        c.fill = header_fill
    for _, row in res_serv.sort_values('demanda', ascending=False).iterrows():
        ws3.append([row['Servicio'], int(row['demanda']),
                    int(row['asignados']), f"{row['tasa_atencion']*100:.1f}%"])
    for col, w in zip('ABCD', [16, 12, 12, 18]):
        ws3.column_dimensions[col].width = w

    # Asignaciones
    ws4 = wb.create_sheet('Asignaciones')
    ws4.append(['Tabla quirúrgica - Planificación MIP'])
    ws4['A1'].font = Font(bold=True, size=12, name='Arial')
    ws4.append([])
    cols = ['ID paciente', 'Servicio', 'Descripción', 'Prioridad', 'Duración (min)',
            'Permanencia_estimada', 'Tipo cama', 'Día', 'Pabellón', 'Hora inicio', 'Hora fin']
    ws4.append(cols)
    for c in ws4[3]:
        c.font = header_font
        c.fill = header_fill
    
    asign_display = asign.copy()
    asign_display['Tipo cama'] = asign_display.apply(
        lambda row: 'UCI' if row['Requiere UCI'] else (
            'Hosp.' if row['Permanencia_estimada'] > 0 else 'Ambulatorio'),
        axis=1
    )
    
    for _, row in asign_display.sort_values(['Día', 'Pabellón', 'Hora inicio (min)']).iterrows():
        ws4.append([row[c] if c in row.index else '' for c in cols])
    for col, w in zip('ABCDEFGHIJK', [12, 14, 38, 11, 14, 18, 13, 7, 10, 12, 12]):
        ws4.column_dimensions[col].width = w

    # No asignados
    ws5 = wb.create_sheet('No asignados')
    ws5.append(['Pacientes no asignados'])
    ws5['A1'].font = Font(bold=True, size=12, name='Arial')
    ws5.append([])
    cols_na = ['ID paciente', 'Servicio', 'Descripción',
               'Duración agendada (min)', 'Prioridad de paciente', 'motivo']
    ws5.append(cols_na)
    for c in ws5[3]:
        c.font = header_font
        c.fill = header_fill
    for _, row in no_asign.iterrows():
        ws5.append([row.get(c, '') for c in cols_na])
    for col, w in zip('ABCDEF', [12, 14, 38, 18, 18, 30]):
        ws5.column_dimensions[col].width = w

    wb.save(ruta_salida)


# ============================================================================
# 7. GENERAR FIGURAS
# ============================================================================

def generar_figuras(asign, ocup, camas, uci, res_prio, carpeta):
    """Genera 4 figuras PNG de análisis."""
    
    plt.rcParams.update({'font.family': 'serif', 'font.size': 10})

    # Fig 1: atención por prioridad
    fig, ax = plt.subplots(figsize=(7, 4))
    prio_orden = res_prio.sort_values('Prioridad')
    x = prio_orden['Prioridad'].astype(int).astype(str)
    ax.bar(x, prio_orden['demanda'], color='#A6C4DE', label='Demanda', edgecolor='black')
    ax.bar(x, prio_orden['asignados'], color='#1F4E78', label='Asignados', edgecolor='black')
    for i, (d, a) in enumerate(zip(prio_orden['demanda'], prio_orden['asignados'])):
        pct = a/d*100 if d else 0
        ax.text(i, d + 8, f'{pct:.0f}%', ha='center', fontsize=9)
    convencion = '5 = mayor' if PRIORIDAD_MAX_ES_MAYOR else '1 = mayor'
    ax.set_xlabel(f'Prioridad clínica ({convencion})')
    ax.set_ylabel('Pacientes')
    ax.set_title('MIP: atención por nivel de prioridad')
    ax.legend()
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta, 'fig_atencion_prioridad.png'), dpi=150)
    plt.close()

    # Fig 2: utilización de pabellones por día
    util = []
    for d in range(DIAS):
        for p in PABELLONES:
            usado = sum(f - i for i, f in ocup[d][p])
            util.append({'Día': d+1, 'Pabellón': p,
                         'Utilización': usado/(HORA_FIN-HORA_INICIO)})
    util_df = pd.DataFrame(util)
    pivot = util_df.pivot(index='Pabellón', columns='Día', values='Utilización')

    fig, ax = plt.subplots(figsize=(8, 4.5))
    im = ax.imshow(pivot.values, cmap='YlOrRd', vmin=0, vmax=1, aspect='auto')
    ax.set_xticks(range(DIAS))
    ax.set_xticklabels([f'Día {d+1}' for d in range(DIAS)])
    ax.set_yticks(range(len(PABELLONES)))
    ax.set_yticklabels([f'Pab. {p}' for p in PABELLONES])
    for i in range(len(PABELLONES)):
        for j in range(DIAS):
            ax.text(j, i, f'{pivot.values[i,j]*100:.0f}%',
                    ha='center', va='center',
                    color='black' if pivot.values[i,j] < 0.6 else 'white',
                    fontsize=9)
    ax.set_title('MIP: utilización de pabellones (%)')
    plt.colorbar(im, ax=ax, label='Utilización')
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta, 'fig_utilizacion_pabellones.png'), dpi=150)
    plt.close()

    # Fig 3: ocupación de camas
    fig, ax = plt.subplots(figsize=(7, 4))
    dias_x = [f'Día {d+1}' for d in range(DIAS)]
    camas_y = [camas.get(d, 0) for d in range(DIAS)]
    ax.bar(dias_x, camas_y, color='#1F4E78',
           label=f'Camas hospitalización (cap. {CAMAS_HOSPITALIZACION})',
           edgecolor='black')
    ax.axhline(CAMAS_HOSPITALIZACION, color='red', linestyle='--', linewidth=1,
               label='Capacidad camas hosp.')
    for i, v in enumerate(camas_y):
        ax.text(i, v + 1, str(v), ha='center', fontsize=9)
    ax.set_ylabel('Camas ocupadas')
    ax.set_title('MIP: ocupación diaria de camas hospitalización')
    ax.legend(loc='lower right')
    ax.grid(axis='y', alpha=0.3)
    ax.set_ylim(0, CAMAS_HOSPITALIZACION + 10)
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta, 'fig_ocupacion_camas.png'), dpi=150)
    plt.close()

    # Fig 4: Gantt día 1
    fig, ax = plt.subplots(figsize=(11, 5))
    colores_serv = {
        'ENT': '#E69F00', 'General': '#56B4E9', 'OBGYN': '#009E73',
        'Ophthalmology': '#F0E442', 'Orthopedics': '#0072B2',
        'Pediatrics': '#D55E00', 'Plastic': '#CC79A7', 'Podiatry': '#999999',
        'Urology': '#882255', 'Vascular': '#117733'
    }
    asign_d1 = asign[asign['Día'] == 1]
    for _, row in asign_d1.iterrows():
        ax.barh(row['Pabellón'], row['Duración (min)'],
                left=row['Hora inicio (min)'],
                color=colores_serv.get(row['Servicio'], '#CCCCCC'),
                edgecolor='black', linewidth=0.5)
    ax.set_yticks(PABELLONES)
    ax.set_yticklabels([f'Pab. {p}' for p in PABELLONES])
    ax.set_xticks(range(8*60, 18*60+1, 60))
    ax.set_xticklabels([f'{h}:00' for h in range(8, 19)])
    ax.set_xlim(8*60-10, 18*60+10)
    ax.invert_yaxis()
    ax.set_xlabel('Hora del día')
    ax.set_title('MIP: tabla quirúrgica - Día 1')
    ax.grid(axis='x', alpha=0.3)
    handles = [mpatches.Patch(color=c, label=s) for s, c in colores_serv.items()]
    ax.legend(handles=handles, bbox_to_anchor=(1.02, 1), loc='upper left',
              fontsize=8, title='Servicio')
    plt.tight_layout()
    plt.savefig(os.path.join(carpeta, 'fig_gantt_dia1.png'),
                dpi=150, bbox_inches='tight')
    plt.close()


# ============================================================================
# 8. EJECUCIÓN PRINCIPAL
# ============================================================================

def main():
    print("=" * 78)
    print("ANÁLISIS DE PLANIFICACIÓN MIP - 7 DÍAS")
    print("ICS2122 - Grupo 11")
    print("=" * 78)

    # Crear carpeta de resultados
    os.makedirs(CARPETA_RESULTADOS, exist_ok=True)
    print(f"\nCarpeta de resultados: {os.path.abspath(CARPETA_RESULTADOS)}")

    # Cargar datos
    print("\n[1] Cargando datos...")
    db = cargar_datos_base()
    lista_espera = cargar_lista_espera()
    planif = cargar_planificacion()
    
    print(f"    Pacientes en lista de espera: {len(lista_espera)}")
    print(f"    Pacientes en planificación MIP: {len(planif)}")

    # Analizar asignaciones
    print("\n[2] Analizando asignaciones...")
    asign, no_asign = analizar_asignaciones(planif, lista_espera)
    print(f"    -> Asignados: {len(asign)}  |  No asignados: {len(no_asign)}")

    # Reconstruir ocupación
    print("\n[3] Reconstruyendo ocupación de pabellones y camas...")
    ocup, camas, uci = reconstruir_ocupacion_y_camas(asign)

    # Calcular KPIs
    print("\n[4] Calculando KPIs:")
    kpis, res_prio, res_serv = calcular_kpis(asign, no_asign, ocup, camas, uci, lista_espera)
    for k, v in kpis.items():
        print(f"    {k:35s}: {v}")

    print("\n[5] Atención por prioridad:")
    print(res_prio.to_string(index=False))

    print("\n[6] Atención por servicio:")
    print(res_serv.to_string(index=False))

    # Exportar Excel
    ruta_xlsx = os.path.join(CARPETA_RESULTADOS, 'resultados_planificacion_MIP.xlsx')
    exportar_excel(asign, no_asign, kpis, res_prio, res_serv, ruta_xlsx)
    print(f"\n[7] Excel guardado: {ruta_xlsx}")

    # Generar figuras
    generar_figuras(asign, ocup, camas, uci, res_prio, CARPETA_RESULTADOS)
    print(f"\n[8] Figuras guardadas en: {CARPETA_RESULTADOS}/")
    print(f"    - fig_atencion_prioridad.png")
    print(f"    - fig_utilizacion_pabellones.png")
    print(f"    - fig_ocupacion_camas.png")
    print(f"    - fig_gantt_dia1.png")

    print("\n" + "=" * 78)
    print("Análisis completado.")
    print("=" * 78)

    return asign, no_asign, kpis


if __name__ == '__main__':
    main()
